# -*- coding: utf-8 -*-
import csv
import os
import pandas as pd
import pandas
from openpyxl import load_workbook

def Generator(fpath):
    xls = pd.read_excel(fpath,sheet_name='actual pre-code')
    xls.to_csv('./tools/Generator/csv/node.csv', index=False)

    with open('./tools/Generator/csv/node.csv', 'rb') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        column = [row[1] for row in reader]

    for num,node in enumerate(column):
        with open('./tools/Generator/csv/node.csv', 'rb') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row['NODE Name'] == node:
                    b = [row['Pre-code']]
                    for element in b:
                        parts = element.split(',')
                        filen = './tools/Generator/plan/NODE %d.xlsx'%(num+1)
                        wb = load_workbook(filename = filen)
                        ws = wb['Case']
                        for i,j in enumerate(parts):
                            col = 'B%s'%(i+2)
                            ws[col] = j
                        wb.save(filen)